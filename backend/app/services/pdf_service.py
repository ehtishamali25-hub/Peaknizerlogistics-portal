import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill


class PDFService:

    OUTPUT_DIR = "uploads/invoices"

    def __init__(self):
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)

    def generate_invoice_pdf(self, invoice_data: dict, rows: list = None) -> str:

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Invoice_{invoice_data['invoice_number']}_{timestamp}.pdf"
        file_path = os.path.join(self.OUTPUT_DIR, filename)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        styles = getSampleStyleSheet()
        story = []

        # ===========================
        # STYLES
        # ===========================
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=10,
            leading=14
        )

        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=8,
            leading=12
        )

        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        )

        title_style = ParagraphStyle(
            'Title',
            parent=styles['Normal'],
            fontSize=14,
            fontName='Helvetica-Bold',
            spaceAfter=10
        )

        right_style = ParagraphStyle(
            'Right',
            parent=normal_style,
            alignment=2
        )

        center_style = ParagraphStyle(
            'Center',
            parent=normal_style,
            alignment=1
        )

        # ===========================
        # HEADER (LEFT + RIGHT SPLIT)
        # ===========================
        company_info = Paragraph("""
        <b>Peaknizer Logistics</b><br/>
        2503D N Harrison St,<br/>
        Arlington, VA 22207<br/>
        Phone: +1 (571) 293-0721<br/>
        Email: contact@peaknizerlogistics.com<br/>
        Website: peaknizerlogistics.com
        """, normal_style)

        invoice_info = Paragraph(f"""
        <b>INVOICE</b><br/>
        Invoice #: {invoice_data['invoice_number']}<br/>
        Issue Date: {invoice_data['issue_date']}<br/>
        Due Date: {invoice_data['due_date']}
        """, right_style)

        header_table = Table([[company_info, invoice_info]], colWidths=[300, 200])
        story.append(header_table)
        story.append(Spacer(1, 20))

        # ===========================
        # CUSTOMER INFO
        # ===========================
        customer_text = f"""
        <b>Bill To:</b><br/>
        {invoice_data['customer_name']}<br/>
        Code: {invoice_data['customer_code']}<br/>
        Status: {invoice_data['status']}
        """
        story.append(Paragraph(customer_text, normal_style))
        story.append(Spacer(1, 20))

        # ===========================
        # TABLE
        # ===========================
        table_data = [["Product Name", "QTY", "Rate", "Amount"]]

        if invoice_data['invoice_type'] == 'shipping':
            table_data.append([
                "Shipping Charges",
                str(invoice_data['quantity']),
                "—",
                f"${invoice_data['total_amount']:.2f}"
            ])
        else:
            table_data.append([
                "Prep Charges",
                str(invoice_data['quantity']),
                f"${invoice_data['rate']:.2f}",
                f"${invoice_data['total_amount']:.2f}"
            ])

        table = Table(table_data, colWidths=[220, 60, 80, 100])

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
        ]))

        story.append(table)
        story.append(Spacer(1, 25))

        # ===========================
        # TERMS (LEFT) + PAYMENT (RIGHT)
        # ===========================
        terms = Paragraph("""
        • All sales are final.<br/>
        • No cancellations, returns, or exchanges.<br/>
        • Damage claims must be made with the carrier.<br/>
        • Company is not responsible after shipment.<br/>
        • No returns without consent.<br/>
        • Customer bears bank/payment charges.
        """, small_style)

        payment = Paragraph(f"""
        <b>PAYMENT</b><br/>
        <b>Balance Due: USD {invoice_data['total_amount']:.2f}</b>
        """, right_style)

        bottom_section = Table([[terms, payment]], colWidths=[320, 180])
        bottom_section.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        story.append(bottom_section)
        story.append(Spacer(1, 40))

        # ===========================
        # BANK DETAILS (CENTER) - DIFFERENT FOR SHIPPING AND PREP
        # ===========================
        if invoice_data['invoice_type'] == 'shipping':
            bank = Paragraph("""
            <b>Qompany LLC</b><br/>
            Bank: Bank of America<br/>
            Routing Number: 051000017<br/>
            Account Number: 435061161139
            """, center_style)
        else:
            bank = Paragraph("""
            <b>Peaknizer LLC</b><br/>
            Bank: Wells Fargo<br/>
            Routing Number: 051400549<br/>
            Account Number: 121000248
            """, center_style)

        story.append(bank)
        story.append(Spacer(1, 15))

        # ===========================
        # FOOTER
        # ===========================
        story.append(Paragraph("Page 1 of 1", center_style))

        doc.build(story)
        return file_path

    # ============================================
    # EXCEL FUNCTION (UNCHANGED)
    # ============================================
    def generate_shipping_details_excel(self, batch_data: dict, rows: list) -> str:

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shipping Details"

        headers = ['Tracking Number', 'Label Cost', 'Customer Name', 'Order Number', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row['tracking_number'])
            ws.cell(row=row_idx, column=2, value=float(row['label_cost']))
            ws.cell(row=row_idx, column=3, value=row['end_customer_name'])
            ws.cell(row=row_idx, column=4, value=row['order_number'])
            ws.cell(row=row_idx, column=5, value=row['date'])

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Shipping_Details_{batch_data['batch_id']}_{timestamp}.xlsx"
        file_path = os.path.join(self.OUTPUT_DIR, filename)

        wb.save(file_path)
        return file_path